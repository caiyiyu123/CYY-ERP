import io
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, status
from pydantic import BaseModel
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.database import get_db
from app.models.commission import CommissionTable, CommissionRate, ShippingTemplate, ShippingRate
from app.models.setting import SystemSetting
from app.utils.deps import require_module

router = APIRouter(tags=["commission-shipping"])

VALID_PLATFORMS = {"wb_local", "wb_cross_border", "ozon_local"}


def _to_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


# ==================== 佣金部分 ====================


@router.post("/api/commission/upload")
def upload_commission(
    platform: str = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_module("commission_shipping")),
):
    if platform not in VALID_PLATFORMS:
        raise HTTPException(status_code=400, detail=f"Invalid platform: {platform}")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files are supported")

    content = file.file.read()
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="File has no data rows")

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data_rows = rows[1:]

    # 删除该平台旧数据
    old_tables = db.query(CommissionTable).filter(CommissionTable.platform == platform).all()
    for t in old_tables:
        db.delete(t)
    db.flush()

    # 创建新表记录
    ct = CommissionTable(platform=platform, filename=file.filename or "")
    db.add(ct)
    db.flush()

    # 解析数据行
    rates = []
    for row in data_rows:
        vals = list(row)
        if not vals or all(v is None for v in vals):
            continue

        if platform == "wb_local":
            category = str(vals[0] or "").strip()
            product_name = str(vals[1] or "").strip()
            rate = _to_float(vals[2]) if len(vals) > 2 else 0.0
            extra = {}
            for i in range(3, min(len(vals), len(headers))):
                extra[headers[i]] = _to_float(vals[i])
            rates.append(CommissionRate(
                table_id=ct.id, category=category, product_name=product_name,
                rate=rate, extra_rates=extra,
            ))

        elif platform == "wb_cross_border":
            category = str(vals[0] or "").strip()
            product_name = str(vals[1] or "").strip()
            rate = _to_float(vals[2]) if len(vals) > 2 else 0.0
            rates.append(CommissionRate(
                table_id=ct.id, category=category, product_name=product_name,
                rate=rate, extra_rates={},
            ))

        elif platform == "ozon_local":
            category = str(vals[0] or "").strip()
            product_name = str(vals[1] or "").strip()
            extra = {}
            for i in range(2, min(len(vals), len(headers))):
                extra[headers[i]] = _to_float(vals[i])
            first_rate = _to_float(vals[2]) if len(vals) > 2 else 0.0
            rates.append(CommissionRate(
                table_id=ct.id, category=category, product_name=product_name,
                rate=first_rate, extra_rates=extra,
            ))

    db.add_all(rates)
    db.commit()
    wb.close()
    return {"detail": f"Uploaded {len(rates)} rows for {platform}", "count": len(rates)}


@router.get("/api/commission/rates")
def list_commission_rates(
    platform: str = Query(...),
    category: str = Query(""),
    product: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _=Depends(require_module("commission_shipping")),
):
    ct = db.query(CommissionTable).filter(CommissionTable.platform == platform).first()
    if not ct:
        return {"rates": [], "headers": [], "total": 0}

    q = db.query(CommissionRate).filter(CommissionRate.table_id == ct.id)
    if category:
        q = q.filter(CommissionRate.category.ilike(f"%{category}%"))
    if product:
        q = q.filter(CommissionRate.product_name.ilike(f"%{product}%"))

    total = q.count()

    # 取第一条提取动态列名
    first = q.first()
    extra_keys = list((first.extra_rates or {}).keys()) if first else []

    rates = q.offset((page - 1) * page_size).limit(page_size).all()

    result = []
    for r in rates:
        item = {
            "id": r.id,
            "category": r.category,
            "product_name": r.product_name,
            "rate": r.rate,
        }
        extras = r.extra_rates or {}
        for k in extra_keys:
            item[k] = extras.get(k, 0.0)
        result.append(item)

    return {"rates": result, "headers": extra_keys, "total": total}


@router.get("/api/commission/info")
def commission_info(
    platform: str = Query(...),
    db: Session = Depends(get_db),
    _=Depends(require_module("commission_shipping")),
):
    ct = db.query(CommissionTable).filter(CommissionTable.platform == platform).first()
    if not ct:
        return {"filename": None, "uploaded_at": None}
    return {"filename": ct.filename, "uploaded_at": ct.uploaded_at.isoformat()}


# ==================== 运费模板部分 ====================


class ShippingRateItem(BaseModel):
    density_min: float
    density_max: float
    price_usd: float


class ShippingTemplateCreate(BaseModel):
    name: str
    date: date
    rates: list[ShippingRateItem]


class ShippingTemplateUpdate(BaseModel):
    name: str
    date: date
    rates: list[ShippingRateItem]


@router.get("/api/shipping/templates")
def list_shipping_templates(
    db: Session = Depends(get_db),
    _=Depends(require_module("commission_shipping")),
):
    templates = db.query(ShippingTemplate).order_by(ShippingTemplate.updated_at.desc()).all()
    result = []
    for t in templates:
        result.append({
            "id": t.id,
            "name": t.name,
            "date": t.date.isoformat(),
            "rate_count": len(t.rates),
            "rates": [
                {"density_min": r.density_min, "density_max": r.density_max, "price_usd": r.price_usd}
                for r in sorted(t.rates, key=lambda x: x.density_min)
            ],
            "created_at": t.created_at.isoformat(),
            "updated_at": t.updated_at.isoformat(),
        })
    return result


@router.post("/api/shipping/templates", status_code=status.HTTP_201_CREATED)
def create_shipping_template(
    data: ShippingTemplateCreate,
    db: Session = Depends(get_db),
    _=Depends(require_module("commission_shipping")),
):
    tpl = ShippingTemplate(name=data.name, date=data.date)
    db.add(tpl)
    db.flush()
    for r in data.rates:
        db.add(ShippingRate(
            template_id=tpl.id,
            density_min=r.density_min,
            density_max=r.density_max,
            price_usd=r.price_usd,
        ))
    db.commit()
    db.refresh(tpl)
    return {"id": tpl.id, "detail": "Template created"}


@router.put("/api/shipping/templates/{tpl_id}")
def update_shipping_template(
    tpl_id: int,
    data: ShippingTemplateUpdate,
    db: Session = Depends(get_db),
    _=Depends(require_module("commission_shipping")),
):
    tpl = db.query(ShippingTemplate).filter(ShippingTemplate.id == tpl_id).first()
    if not tpl:
        raise HTTPException(status_code=404, detail="Template not found")
    tpl.name = data.name
    tpl.date = data.date
    # 删除旧运费明细，重建
    db.query(ShippingRate).filter(ShippingRate.template_id == tpl.id).delete()
    for r in data.rates:
        db.add(ShippingRate(
            template_id=tpl.id,
            density_min=r.density_min,
            density_max=r.density_max,
            price_usd=r.price_usd,
        ))
    db.commit()
    return {"detail": "Template updated"}


@router.delete("/api/shipping/templates/{tpl_id}")
def delete_shipping_template(
    tpl_id: int,
    db: Session = Depends(get_db),
    _=Depends(require_module("commission_shipping")),
):
    tpl = db.query(ShippingTemplate).filter(ShippingTemplate.id == tpl_id).first()
    if not tpl:
        raise HTTPException(status_code=404, detail="Template not found")
    db.delete(tpl)
    # 如果删除的是默认模板，清除默认设置
    s = db.query(SystemSetting).filter(SystemSetting.key == "default_shipping_template").first()
    if s and s.value == str(tpl_id):
        db.delete(s)
    db.commit()
    return {"detail": "Template deleted"}


@router.get("/api/shipping/default-template")
def get_default_template(
    db: Session = Depends(get_db),
    _=Depends(require_module("commission_shipping")),
):
    s = db.query(SystemSetting).filter(SystemSetting.key == "default_shipping_template").first()
    return {"id": int(s.value) if s else None}


@router.put("/api/shipping/default-template/{tpl_id}")
def set_default_template(
    tpl_id: int,
    db: Session = Depends(get_db),
    _=Depends(require_module("commission_shipping")),
):
    tpl = db.query(ShippingTemplate).filter(ShippingTemplate.id == tpl_id).first()
    if not tpl:
        raise HTTPException(status_code=404, detail="Template not found")
    s = db.query(SystemSetting).filter(SystemSetting.key == "default_shipping_template").first()
    if s:
        s.value = str(tpl_id)
    else:
        db.add(SystemSetting(key="default_shipping_template", value=str(tpl_id)))
    db.commit()
    return {"id": tpl_id}
